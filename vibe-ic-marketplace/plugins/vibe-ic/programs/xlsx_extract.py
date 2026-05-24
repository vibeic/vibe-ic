#!/usr/bin/env python3
"""
xlsx_extract.py — Extract tables from .xlsx spec documents into JSON.

Closes the "plugin ignores the xlsx in input/docs/" gap found 2026-04-22:
vendor specs often ship protocol commands, CRC golden values, OTP memory
contents, pin-mux tables, and timing parameters in .xlsx form. Doc-gen
skills (cmd-protocol-gen, regmap-gen, rtl-constants-gen) previously only
read .pdf / .txt / .md, missing every spreadsheet table.

Output
------
One JSON object per sheet:
  {
    "path": "input/docs/CMD整理.xlsx",
    "sheets": {
      "Sheet1": {
        "header": ["CMD", "Payload", "Response", "CRC"],
        "rows":   [ ["70", "00 00 3D", "71", "0x3D"], ... ],
        "row_count": 16,
      },
      ...
    }
  }

No interpretation — just extraction. The calling skill decides how to
map rows onto L3/L4/L8. Designed to be idempotent (same xlsx → same
JSON) so a provenance hash can be computed.

Usage
-----
    xlsx_extract.py <file.xlsx> [--json out.json]
    xlsx_extract.py <dir>  [--json out.json]       # scans dir for *.xlsx

Exit codes
----------
    0 = extracted at least one sheet with content
    1 = no xlsx found / all empty
    2 = io error

Dependencies
------------
    openpyxl (fallback to zipfile+xml parse if openpyxl missing)
"""
from __future__ import annotations

import argparse
import json
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import List, Dict, Any


def _try_openpyxl(path: Path) -> Dict[str, Any] | None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return {"error": f"openpyxl load failed: {exc}"}
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            # Convert None → "" for cleaner JSON
            rows.append([("" if c is None else str(c)) for c in row])
        # Trim trailing all-empty rows
        while rows and all(c == "" for c in rows[-1]):
            rows.pop()
        if not rows:
            continue
        header = rows[0]
        data_rows = rows[1:]
        sheets[name] = {
            "header": header,
            "rows": data_rows,
            "row_count": len(data_rows),
        }
    wb.close()
    return {"sheets": sheets}


def _zip_fallback(path: Path) -> Dict[str, Any]:
    """Minimal XLSX parser using only stdlib (zipfile + ElementTree).
    Extracts cell values from each sheet. Does not resolve shared strings
    for styled cells — sufficient for tables of plain text/numbers."""
    ns = {"": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    sheets: Dict[str, Dict[str, Any]] = {}
    try:
        with zipfile.ZipFile(path) as z:
            # Load shared strings
            strings: List[str] = []
            if "xl/sharedStrings.xml" in z.namelist():
                tree = ET.fromstring(z.read("xl/sharedStrings.xml"))
                for si in tree.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si"):
                    # Concatenate all <t> descendants
                    t = "".join(t.text or ""
                                for t in si.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"))
                    strings.append(t)
            # Find sheet names from workbook.xml
            wb_tree = ET.fromstring(z.read("xl/workbook.xml"))
            sheet_names = []
            for s in wb_tree.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"):
                sheet_names.append(s.attrib.get("name", f"Sheet{len(sheet_names)+1}"))

            sheet_paths = sorted([n for n in z.namelist()
                                   if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")])
            for i, sheet_path in enumerate(sheet_paths):
                name = sheet_names[i] if i < len(sheet_names) else f"Sheet{i+1}"
                tree = ET.fromstring(z.read(sheet_path))
                rows: List[List[str]] = []
                # Collect max column across all rows to pad
                for row_el in tree.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row"):
                    row_vals: Dict[int, str] = {}
                    for c in row_el.findall("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c"):
                        ref = c.attrib.get("r", "")
                        # Extract column letter(s)
                        col_letters = "".join(ch for ch in ref if ch.isalpha())
                        col_idx = 0
                        for ch in col_letters:
                            col_idx = col_idx * 26 + (ord(ch.upper()) - ord('A') + 1)
                        col_idx -= 1
                        t = c.attrib.get("t", "")
                        v = c.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
                        is_el = c.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}is")
                        val = ""
                        if t == "s" and v is not None and v.text is not None:
                            idx = int(v.text)
                            if 0 <= idx < len(strings):
                                val = strings[idx]
                        elif t == "inlineStr" and is_el is not None:
                            val = "".join(sub.text or "" for sub in is_el.iter()
                                          if sub.tag.endswith("}t"))
                        elif v is not None and v.text is not None:
                            val = v.text
                        row_vals[col_idx] = val
                    if row_vals:
                        max_col = max(row_vals.keys())
                        rows.append([row_vals.get(i, "") for i in range(max_col + 1)])
                # Normalize row widths
                if rows:
                    width = max(len(r) for r in rows)
                    rows = [r + [""] * (width - len(r)) for r in rows]
                    while rows and all(c == "" for c in rows[-1]):
                        rows.pop()
                if rows:
                    sheets[name] = {
                        "header": rows[0],
                        "rows": rows[1:],
                        "row_count": len(rows) - 1,
                    }
    except Exception as exc:
        return {"error": f"zip-fallback parse failed: {exc}"}
    return {"sheets": sheets}


def extract_xlsx(path: Path) -> Dict[str, Any]:
    """Extract one .xlsx. Uses openpyxl if available, else stdlib fallback."""
    result: Dict[str, Any] = {"path": str(path), "size_bytes": path.stat().st_size}
    r = _try_openpyxl(path)
    if r is None or "error" in (r or {}):
        r = _zip_fallback(path)
    result.update(r)
    return result


def main(argv: List[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    p.add_argument("target", help="xlsx file OR directory to scan")
    p.add_argument("--json", help="Write combined JSON report to this path")
    p.add_argument("--min-rows", type=int, default=1,
                   help="Minimum data rows required per sheet to be non-empty")
    args = p.parse_args(argv)

    tp = Path(args.target).resolve()
    files: List[Path] = []
    if tp.is_file() and tp.suffix.lower() == ".xlsx":
        files = [tp]
    elif tp.is_dir():
        files = sorted(tp.rglob("*.xlsx"))
    else:
        print(f"xlsx_extract: not a file or dir: {tp}", file=sys.stderr)
        return 2

    if not files:
        print("xlsx_extract: no .xlsx files found")
        return 1

    report = {"files": []}
    any_content = False
    for f in files:
        entry = extract_xlsx(f)
        # Count total rows
        total_rows = sum(s.get("row_count", 0) for s in entry.get("sheets", {}).values())
        entry["total_data_rows"] = total_rows
        if total_rows >= args.min_rows:
            any_content = True
        report["files"].append(entry)

    # Print per-file summary
    print(f"\n=== xlsx_extract ===")
    for e in report["files"]:
        sheets = e.get("sheets", {})
        print(f"  {e['path']}")
        if "error" in e:
            print(f"    ERROR: {e['error']}")
            continue
        for sname, s in sheets.items():
            header_str = ", ".join(s['header'][:5]) + ("..." if len(s['header']) > 5 else "")
            print(f"    [{sname}] {s['row_count']} rows, cols: {header_str}")

    if args.json:
        Path(args.json).parent.mkdir(parents=True, exist_ok=True)
        Path(args.json).write_text(json.dumps(report, ensure_ascii=False, indent=2))

    return 0 if any_content else 1


if __name__ == "__main__":
    sys.exit(main())
