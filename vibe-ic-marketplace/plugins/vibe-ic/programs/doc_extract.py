#!/usr/bin/env python3
"""
doc_extract.py — Convert vendor docs (.doc/.docx/.pdf/.pptx/.xlsx/.txt) to
plain-text or structured JSON for downstream Phase 1 (doc-extraction) / spec-derivation skills.

Every benchmark project's `input/docs/` folder ships a different mix of
vendor doc formats. The clean-slate fresh-agent run revealed there was no
plugin skill to handle this, so every agent rebuilds the same shell pipeline
(pdftotext + libreoffice + openpyxl). This script unifies them.

Supported inputs:
  .pdf       → pdftotext (poppler)
  .doc/.docx → libreoffice headless --convert-to txt:Text
  .ppt/.pptx → libreoffice headless --convert-to txt:Text
  .xls/.xlsx → openpyxl (workbook → JSON {sheet_name: [[row...]]} + flat text)
  .txt/.md   → cat (UTF-8 normalize)
  .html/.htm → python -m html2text (best effort; falls back to strip-tags)

Usage:
    python3 doc_extract.py --in-dir ./input/docs --out-dir ./generated_docs
    # Per-file:
    python3 doc_extract.py --in-file ./input/docs/spec.pdf --out-dir /tmp

Outputs in --out-dir:
    <stem>.txt        — plain text extraction
    <stem>.json       — for xlsx, structured cell data; for others, metadata
    INDEX.json        — manifest: {input_path, output_text, char_count, status}

Exit codes: 0 = at least one doc extracted, 1 = no docs / all failed, 2 = usage.

Generality: domain-agnostic. Works for any vendor's documentation pack.
"""
from __future__ import annotations
import argparse, json, re, shutil, subprocess, sys
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import List, Optional


@dataclass
class ExtractResult:
    input_path: str
    format: str
    output_text: str
    output_json: Optional[str]
    char_count: int
    status: str
    error: str = ""
    # v0.119.33: capture file_size + coverage_score (text_chars /
    # file_size_bytes) so downstream gates can flag figure-heavy
    # PDFs whose `pdftotext` output is essentially empty. See
    # binary_doc_low_extraction_warn (LL-36).
    file_size: int = 0
    coverage_score: float = 0.0
    extractor_warnings: List[str] = field(default_factory=list)


def _which(name: str) -> Optional[str]:
    return shutil.which(name)


def extract_pdf(p: Path, out_txt: Path) -> ExtractResult:
    if not _which("pdftotext"):
        return ExtractResult(str(p), "pdf", "", None, 0, "FAIL", "pdftotext not in PATH")
    try:
        subprocess.run(["pdftotext", "-layout", str(p), str(out_txt)],
                       check=True, capture_output=True, timeout=120)
        text = out_txt.read_text(errors="replace") if out_txt.exists() else ""
        return ExtractResult(str(p), "pdf", str(out_txt), None, len(text), "PASS")
    except Exception as e:
        return ExtractResult(str(p), "pdf", "", None, 0, "FAIL", str(e))


def _extract_ooxml_zip_text(p: Path, slide_or_doc: str) -> str:
    """stdlib fallback for .pptx / .docx — read OOXML zip + extract <a:t> / <w:t> text.
    Used when libreoffice silently fails (which it does on some PPTX builds with
    SfxBaseModel impl_store 0xc10 errors).

    v0.76.3: pptx pattern now also matches notesSlides/notesSlide*.xml so
    speaker-note text (where many vendors put spec details) survives extraction."""
    import zipfile
    from xml.etree import ElementTree as ET
    chunks: List[str] = []
    if slide_or_doc == "pptx":
        ns = "http://schemas.openxmlformats.org/drawingml/2006/main"
        member_pat = re.compile(r"^ppt/(?:slides/slide|notesSlides/notesSlide)\d+\.xml$")
        tag = f"{{{ns}}}t"
    else:  # docx
        ns = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
        member_pat = re.compile(r"^word/document\.xml$")
        tag = f"{{{ns}}}t"
    with zipfile.ZipFile(p) as z:
        for name in sorted(z.namelist()):
            if not member_pat.match(name):
                continue
            try:
                root = ET.parse(z.open(name)).getroot()
            except ET.ParseError:
                continue
            for el in root.iter(tag):
                if el.text and el.text.strip():
                    chunks.append(el.text)
    return "\n".join(chunks)


def extract_office_libreoffice(p: Path, out_dir: Path, fmt: str) -> ExtractResult:
    """v0.76.3: 3-tier fallback with per-tier error accumulation —
       Tier 1: stdlib zip+xml read for .pptx/.docx (most reliable, no deps).
       Tier 2: libreoffice headless --convert-to txt:Text (ASCII tmp stem
               for CJK filename safety).
       Tier 3: libreoffice with isolated UserInstallation profile (avoids
               clashes with another running libreoffice instance).

    All tiers' failure reasons are accumulated into the final FAIL message
    so a generic 'all tiers failed' line no longer hides the actual root
    cause (regression vs v0.76.2)."""
    final_txt = out_dir / (p.stem + ".txt")
    suf = p.suffix.lower()
    errs: List[str] = []  # per-tier diagnostic, joined into final FAIL message

    # Tier 1: stdlib OOXML reader for .pptx/.docx
    if suf in (".pptx", ".docx"):
        kind = "pptx" if suf == ".pptx" else "docx"
        try:
            text = _extract_ooxml_zip_text(p, kind)
            if text.strip():
                final_txt.write_text(text, encoding="utf-8")
                return ExtractResult(str(p), fmt, str(final_txt), None, len(text), "PASS")
            errs.append("tier1: empty extraction (no <a:t>/<w:t> text in zip)")
        except Exception as e:
            errs.append(f"tier1: {e.__class__.__name__}: {e}")

    if not _which("libreoffice"):
        errs.append("libreoffice not in PATH")
        return ExtractResult(str(p), fmt, "", None, 0, "FAIL", "; ".join(errs))

    import tempfile, shutil
    # Tier 2: libreoffice in tmp dir with ASCII stem (handles CJK filenames)
    try:
        with tempfile.TemporaryDirectory(prefix="lo_extract_") as td:
            tdp = Path(td)
            ascii_stem = "doc_" + str(abs(hash(p.name)))[:12]
            ascii_in = tdp / (ascii_stem + p.suffix)
            shutil.copy2(p, ascii_in)
            r2 = subprocess.run([
                "libreoffice", "--headless", "--convert-to", "txt:Text",
                "--outdir", str(tdp), str(ascii_in),
            ], capture_output=True, timeout=180)
            cands = sorted(tdp.glob("*.txt"))
            if cands and cands[0].stat().st_size > 0:
                shutil.copy2(cands[0], final_txt)
                text = final_txt.read_text(errors="replace")
                return ExtractResult(str(p), fmt, str(final_txt), None, len(text), "PASS")
            errs.append(f"tier2: libreoffice exit={r2.returncode}, no/empty .txt")
    except Exception as e:
        errs.append(f"tier2: {e.__class__.__name__}: {e}")

    # Tier 3: libreoffice with isolated profile
    try:
        with tempfile.TemporaryDirectory(prefix="lo_profile_") as profile_td, \
             tempfile.TemporaryDirectory(prefix="lo_out_") as out_td:
            tdp = Path(out_td)
            ascii_stem = "doc_" + str(abs(hash(p.name)))[:12]
            ascii_in = tdp / (ascii_stem + p.suffix)
            shutil.copy2(p, ascii_in)
            r3 = subprocess.run([
                "libreoffice", "--headless", f"-env:UserInstallation=file://{profile_td}",
                "--convert-to", "txt:Text",
                "--outdir", str(tdp), str(ascii_in),
            ], capture_output=True, timeout=180)
            cands = sorted(tdp.glob("*.txt"))
            if cands and cands[0].stat().st_size > 0:
                shutil.copy2(cands[0], final_txt)
                text = final_txt.read_text(errors="replace")
                return ExtractResult(str(p), fmt, str(final_txt), None, len(text), "PASS")
            errs.append(f"tier3: libreoffice exit={r3.returncode}, no/empty .txt")
    except Exception as e:
        errs.append(f"tier3: {e.__class__.__name__}: {e}")

    return ExtractResult(str(p), fmt, "", None, 0, "FAIL", "; ".join(errs))


def extract_xlsx(p: Path, out_dir: Path) -> ExtractResult:
    try:
        import openpyxl  # noqa
    except Exception:
        return ExtractResult(str(p), "xlsx", "", None, 0, "FAIL", "openpyxl not installed (pip install openpyxl)")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(p, data_only=True, read_only=True)
        struct = {}
        flat_lines: List[str] = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_clean = ["" if v is None else str(v) for v in row]
                if any(c.strip() for c in row_clean):
                    rows.append(row_clean)
                    flat_lines.append(f"[{sn}] " + "\t".join(row_clean))
            struct[sn] = rows
        out_txt = out_dir / (p.stem + ".txt")
        out_json = out_dir / (p.stem + ".json")
        out_txt.write_text("\n".join(flat_lines), encoding="utf-8")
        out_json.write_text(json.dumps(struct, ensure_ascii=False, indent=2))
        return ExtractResult(str(p), "xlsx", str(out_txt), str(out_json), len("\n".join(flat_lines)), "PASS")
    except Exception as e:
        return ExtractResult(str(p), "xlsx", "", None, 0, "FAIL", str(e))


def extract_text(p: Path, out_txt: Path) -> ExtractResult:
    try:
        text = p.read_text(errors="replace")
        out_txt.write_text(text, encoding="utf-8")
        return ExtractResult(str(p), "text", str(out_txt), None, len(text), "PASS")
    except Exception as e:
        return ExtractResult(str(p), "text", "", None, 0, "FAIL", str(e))


def extract_html(p: Path, out_txt: Path) -> ExtractResult:
    try:
        text = p.read_text(errors="replace")
        text = re.sub(r'<[^>]+>', ' ', text)  # strip tags (best-effort)
        text = re.sub(r'\s+', ' ', text).strip()
        out_txt.write_text(text, encoding="utf-8")
        return ExtractResult(str(p), "html", str(out_txt), None, len(text), "PASS")
    except Exception as e:
        return ExtractResult(str(p), "html", "", None, 0, "FAIL", str(e))


_FMT_DISPATCH = {
    ".pdf": ("pdf", "pdf"),
    ".doc": ("doc", "office"),
    ".docx": ("docx", "office"),
    ".ppt": ("ppt", "office"),
    ".pptx": ("pptx", "office"),
    ".xls": ("xls", "office"),
    ".xlsx": ("xlsx", "xlsx"),
    ".txt": ("text", "text"),
    ".md": ("text", "text"),
    ".html": ("html", "html"),
    ".htm": ("html", "html"),
}


def extract_one(p: Path, out_dir: Path) -> ExtractResult:
    suf = p.suffix.lower()
    if suf not in _FMT_DISPATCH:
        return ExtractResult(str(p), "skipped", "", None, 0, "SKIP", f"unsupported suffix: {suf}")
    fmt, dispatch = _FMT_DISPATCH[suf]
    out_txt = out_dir / (p.stem + ".txt")
    if dispatch == "pdf":
        return extract_pdf(p, out_txt)
    if dispatch == "office":
        return extract_office_libreoffice(p, out_dir, fmt)
    if dispatch == "xlsx":
        return extract_xlsx(p, out_dir)
    if dispatch == "text":
        return extract_text(p, out_txt)
    if dispatch == "html":
        return extract_html(p, out_txt)
    return ExtractResult(str(p), "unknown", "", None, 0, "FAIL", "no dispatch matched")


def main(argv: List[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.strip().split("\n")[0])
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--in-dir", type=Path)
    g.add_argument("--in-file", type=Path)
    ap.add_argument("--out-dir", type=Path, required=True)
    ap.add_argument("--recurse", action="store_true", help="(default) recurse subdirs")
    ap.add_argument("--no-recurse", dest="recurse", action="store_false")
    ap.set_defaults(recurse=True)
    args = ap.parse_args(argv)
    args.out_dir.mkdir(parents=True, exist_ok=True)

    inputs: List[Path] = []
    if args.in_file:
        if not args.in_file.is_file():
            print(f"ERROR: not a file: {args.in_file}", file=sys.stderr)
            return 2
        inputs = [args.in_file]
    else:
        if not args.in_dir.is_dir():
            print(f"ERROR: not a dir: {args.in_dir}", file=sys.stderr)
            return 2
        glob = "**/*" if args.recurse else "*"
        inputs = [p for p in args.in_dir.glob(glob) if p.is_file()]

    results: List[ExtractResult] = []
    for p in sorted(inputs):
        r = extract_one(p, args.out_dir)
        # v0.119.33: enrich with coverage_score for binary-doc
        # diagnostics (LL-36 / P1.4). Skip on stat() failure.
        try:
            sz = p.stat().st_size
            r.file_size = sz
            if sz > 0:
                r.coverage_score = round(r.char_count / sz, 6)
        except OSError:
            pass
        # v0.119.33: figure-heavy PDF hint when fallback extractors
        # aren't available. Threshold (0.02) chosen to match
        # binary_doc_low_extraction_warn so the two stay aligned.
        if (p.suffix.lower() == ".pdf" and r.status == "PASS"
                and r.coverage_score and r.coverage_score < 0.02):
            try:
                import importlib  # noqa
                have_plumber = importlib.util.find_spec("pdfplumber") is not None
                have_fitz = importlib.util.find_spec("fitz") is not None
            except Exception:
                have_plumber = have_fitz = False
            if not (have_plumber or have_fitz):
                r.extractor_warnings.append(
                    f"coverage_score {r.coverage_score:.3f} — "
                    "install pdfplumber/pymupdf for figure-heavy "
                    "PDF fallback extraction"
                )
        results.append(r)
        print(f"[{r.status}] {r.format:6s} {p}  ({r.char_count} chars)" + (f"  err={r.error}" if r.error else ""))

    index = args.out_dir / "INDEX.json"
    index.write_text(json.dumps([asdict(r) for r in results], ensure_ascii=False, indent=2))
    pass_count = sum(1 for r in results if r.status == "PASS")
    fail_count = sum(1 for r in results if r.status == "FAIL")
    skip_count = sum(1 for r in results if r.status == "SKIP")
    print(f"---\n{pass_count} PASS / {fail_count} FAIL / {skip_count} SKIP. INDEX: {index}")
    return 0 if pass_count > 0 and fail_count == 0 else (0 if pass_count > 0 else 1)


if __name__ == "__main__":
    sys.exit(main())
